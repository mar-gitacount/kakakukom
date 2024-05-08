from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
import sys

num = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# ここで num を使用して何かを実行する
(f"受け取った引数: {num}")
url = "https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_Spec106=1,2&pdf_vi=c"
url = f"https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090&pdf_Spec106=1,2,4&pdf_vi=c&pdf_pg={num}"


# ?モデル、ブレスレット、文字盤を抽出する関数
def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    beltpattern = r"\[(.*?)\]|\((.*?)\)"
    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。
    items = {"model": model, "beltmatches": beltmatches}
    return items


# ?エクセルに入力する関数
def wsinsert(values, sheet):
    print("wsinsert関数", values)
    sheet.append(values)
    # for item in values:
    #     sheet.append(item)


# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
# ファイル名に日付を組み込む

file_name = f"output_{today_date}.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(
        [
            "モデル名",
            "リファレンスNO",
            "ブレスレット",
            "新品価格",
            "中古価格",
            "順位",
            "URL",
        ]
    )
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active


# SeleniumのWebDriverを初期化
driver = webdriver.Chrome()  # または他のブラウザに合わせて選択

# URLを開く
driver.get(url)

# Seleniumがページのロードを待つなどの適切な待機処理が必要な場合はここで実施

# ページのHTMLを取得
page_source = driver.page_source

# BeautifulSoupを使ってHTMLを解析
soup = BeautifulSoup(page_source, "html.parser")


# !ここから処理スタート

# <tbody> タグ内のテキストを抽出して表示
tbody_tag = soup.find("body")
# tbody_tag = soup.find("tbody")
table_get = tbody_tag.find("table", id="compTblList")
# print(table_get)
tr_get = table_get.find_all("tr", class_="tr-border")

td_get = table_get.find_all("td", class_="end")
td_get = table_get.find_all("a", class_="ckitanker")
# 取得するアイテム一覧アクセスする
for item in td_get:
    # td_get = item.find("td", class_="end")
    # print(item)
    # 各アイテムごとのurl
    # アイテム配列
    # ?変数初期化する
    itemname = ""
    price = ""
    usedprice = ""
    ranking = ""

    href = item.get("href")
    print(href)
    driver.get(href)
    item_page_source = driver.page_source
    item_soup = BeautifulSoup(item_page_source, "html.parser")
    # bodytag = item_soup.find("div", id="watch-accessory")
    itemboxbottom = item_soup.find("div", class_="itmBoxBottom")
    # アイテムネーム
    # 正規表現する必要がある。
    itemname = item_soup.find("div", id="titleBox").find("h2").get_text(strip=True)
    print("テキスト", itemname)
    # リファレンスナンバー取得する。
    ref_pattern = r"\b(\d{4,6})([a-zA-Z]+)?\b"
    refmatches = re.findall(
        ref_pattern,
        itemname,
        flags=re.UNICODE,
    )

    refmatches = "".join(["".join(t) for t in refmatches])
    # リファレンスナンバーを取り除く。
    modelmatches = re.sub(ref_pattern, "", itemname)
    # モデル名、ベルト、文字盤を抽出する
    # 辞書型配列が返ってきてる。
    # モデル名を抽出する

    model_belt_bracelet_item = model_validete_imput(modelmatches)
    print("モデル、ベルト、文字盤の連想配列", model_belt_bracelet_item)
    brlt_brecelet = "".join(t[0] for t in model_belt_bracelet_item["beltmatches"])

    # 前後に空白がある表現をすべて取得する
    empty_pattern = r"\s+\b(\w+)\b\s+"
    modelmatches = re.findall(
        empty_pattern,
        modelmatches,
        flags=re.UNICODE,
    )

    # 新品値段
    try:
        price = item_soup.find("span", class_="priceTxt").get_text(strip=True)
        print("値段", price)
    except AttributeError:
        print("値段なしからの文字列を代入する")

    # 中古値段
    try:
        usedpriceelement = item_soup.find("span", class_="usedPriceTxt")
        if usedpriceelement is not None:
            usedprice = usedpriceelement.get_text(strip=True)
        else:
            usedprice = ""
        print("中古", usedprice)
    except ArithmeticError:
        print("中古の値段なし")

    # 順位
    try:
        rankingparent = item_soup.find("div", id="ovBtnBox")
        ranking = rankingparent.find("span", class_="num").get_text(strip=True)
        print("順位", ranking)
    except ArithmeticError:
        print("ランキングなし")
    print("--------------------------")
    insertitems = [
        model_belt_bracelet_item["model"],
        refmatches,
        brlt_brecelet,
        price,
        usedprice,
        ranking,
        href,
    ]
    wsinsert(insertitems, ws)

wb.save(file_name)

# !ここまで
